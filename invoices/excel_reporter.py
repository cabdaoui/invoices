# invoices/excel_reporter.py
from __future__ import annotations

from pathlib import Path
from typing import Iterable, Mapping, Any, Sequence, Optional
import os

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from invoices.utils import load_env_config  # <-- pour lire OUTPUT_DIR depuis env.json

# Ordre des colonnes attendu par pdf_parser.extract_invoice_data()
COLUMNS: Sequence[str] = (
    "fichier",         # nom du PDF
    "facture",         # numéro de facture
    "date",            # dd/mm/yyyy
    "total_ttc",       # 255,63€ (converti en nombre si possible)
    "periode",         # ex. "Octobre 2025" si détecté
    "source_montant",  # info debug : "TTC* mois" / "TTC* générique" / "montant générique"
)

ALLOWED_EXCEL_NAME = "invoices_extract.xlsx"


# ---------------------------
# Helpers chemins & fichiers
# ---------------------------

def _project_root() -> Path:
    """Racine du projet: parent de invoices/ ou WORKSPACE Jenkins si défini."""
    package_dir = Path(__file__).resolve().parent  # .../invoices
    ws = os.environ.get("WORKSPACE")
    if ws:
        return Path(ws).resolve()
    return package_dir.parent  # .../invoices_project

def _resolve_dir(base: Path, value: str) -> Path:
    p = Path(value)
    return (base / p).resolve() if not p.is_absolute() else p.resolve()

def _ensure_parent_dir(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


# ---------------------------
# Parsing / formats
# ---------------------------

def _number_from_amount(val: Any) -> Optional[float]:
    """
    Essaie de convertir un montant texte ('1 234,56€', '1,234.56', '1234.56€') en float.
    Retourne None si non convertible (on laissera alors la valeur texte telle quelle).
    """
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # retire les devises courantes et espaces
    s = s.replace("€", "").replace("$", "").replace(" ", "")
    # heuristique du séparateur décimal
    if "," in s and "." in s:
        if s.rfind(".") > s.rfind(","):
            # "1,234.56" -> milliers: ',', décimal: '.'
            s = s.replace(",", "")
        else:
            # "1.234,56" -> milliers: '.', décimal: ','
            s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        # "1234,56" (FR)
        s = s.replace(",", ".")
    # else: "1234.56" (US) ou entier
    try:
        return float(s)
    except Exception:
        return None


# ---------------------------
# Mise en forme Excel
# ---------------------------

def _apply_table_style(ws: Worksheet, n_rows: int, n_cols: int) -> None:
    # Style en-tête
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # En-têtes (ligne 1)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        cell.border = border

    # Corps
    for r in range(2, n_rows + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border

    # Figer la première ligne et activer l’autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Ajustement simple des largeurs de colonnes
    for idx in range(1, n_cols + 1):
        letter = get_column_letter(idx)
        max_len = 0
        for r in range(1, n_rows + 1):
            v = ws.cell(row=r, column=idx).value
            l = len(str(v)) if v is not None else 0
            max_len = max(max_len, l)
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), 60)


# ---------------------------
# Écriture Excel
# ---------------------------

def write_report(
    rows: Iterable[Mapping[str, Any]],
    xlsx_path: str | Path,
    sheet_name: str = "Reporting"
) -> Path:
    """
    Écrit un Excel de reporting aligné sur pdf_parser.extract_invoice_data().
    - rows: itérable de dicts ayant idéalement les clés COLUMNS
    - xlsx_path: chemin cible du fichier .xlsx (écrasé si existe)
    - sheet_name: nom de l’onglet
    Retourne le chemin absolu du fichier .xlsx généré.
    """
    xlsx_path = Path(xlsx_path).resolve()
    _ensure_parent_dir(xlsx_path)

    wb = Workbook()
    ws = wb.active
    # Sanitize minimal du nom d’onglet (Excel limite à 31 chars; interdit: []:*?/\\)
    clean_title = "".join(ch for ch in sheet_name if ch not in '[]:*?/\\')
    ws.title = clean_title[:31] if clean_title else "Reporting"

    # En-têtes
    ws.append(list(COLUMNS))

    # Lignes
    count = 0
    for r in rows:
        count += 1
        fichier = r.get("fichier", "")
        facture = r.get("facture", "")
        date = r.get("date", "")
        total_ttc_raw = r.get("total_ttc", "")
        periode = r.get("periode", "")
        source_montant = r.get("source_montant", "")

        # Conversion du total_ttc si possible
        total_number = _number_from_amount(total_ttc_raw)

        ws.append([
            fichier,
            facture,
            date,
            total_number if total_number is not None else total_ttc_raw,
            periode,
            source_montant
        ])

    # Format nombre à 2 décimales (FR-like) pour la colonne total_ttc quand numérique
    for row_idx in range(2, count + 2):  # Colonne 4 = total_ttc
        cell = ws.cell(row=row_idx, column=4)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'

    _apply_table_style(ws, n_rows=count + 1, n_cols=len(COLUMNS))
    wb.save(xlsx_path)
    return xlsx_path


# ---------------------------
# Fonctions ajoutées pour pipeline
# ---------------------------

def default_output_xlsx_path() -> Path:
    """Construit <OUTPUT_DIR>/invoices_extract.xlsx (dossier créé si besoin)."""
    env = load_env_config()
    root = _project_root()
    output_dir = _resolve_dir(root, env.get("OUTPUT_DIR", "./output"))
    output_dir.mkdir(parents=True, exist_ok=True)
    return (output_dir / ALLOWED_EXCEL_NAME).resolve()

def write_report_to_output(
    rows: Iterable[Mapping[str, Any]],
    sheet_name: str = "Reporting"
) -> Path:
    """
    Écrit le reporting dans <OUTPUT_DIR>/invoices_extract.xlsx (nom imposé).
    Retourne le chemin absolu.
    """
    xlsx_path = default_output_xlsx_path()
    return write_report(rows, xlsx_path, sheet_name)
